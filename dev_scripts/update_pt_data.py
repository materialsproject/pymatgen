"""Developer script to convert YAML periodic table to JSON format.
Created on 2011-11-15."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from itertools import product

import requests
from monty.dev import requires
from monty.serialization import dumpfn, loadfn
from ruamel import yaml

from pymatgen.core import Element, get_el_sp

try:
    from bs4 import BeautifulSoup
except ImportError:
    BeautifulSoup = None

ptable_yaml_path = "periodic_table.yaml"


def parse_oxi_state():
    data = loadfn(ptable_yaml_path)
    with open("oxidation_states.txt") as file:
        oxi_data = file.read()
    oxi_data = re.sub("[\n\r]", "", oxi_data)
    patt = re.compile("<tr>(.*?)</tr>", re.MULTILINE)

    for match in patt.finditer(oxi_data):
        line = match[1]
        line = re.sub("</td>", "", line)
        line = re.sub("(<td>)+", "<td>", line)
        line = re.sub("</*a[^>]*>", "", line)
        el = None
        oxi_states = []
        common_oxi = []
        for tok in re.split("<td>", line.strip()):
            match2 = re.match(r"<b>([A-Z][a-z]*)</b>", tok)
            if match2:
                el = match2[1]
            else:
                match3 = re.match(r"(<b>)*([\+\-]\d)(</b>)*", tok)
                if match3:
                    oxi_states += [int(match3[2])]
                    if match3[1]:
                        common_oxi += [int(match3[2])]
        if el in data:
            del data[el]["Max oxidation state"]
            del data[el]["Min oxidation state"]
            del data[el]["Oxidation_states"]
            del data[el]["Common_oxidation_states"]
            data[el]["Oxidation states"] = oxi_states
            data[el]["Common oxidation states"] = common_oxi
        else:
            print(el)
    with open("periodic_table2.yaml", mode="w") as file:
        yaml.dump(data, file)


def parse_ionic_radii():
    data = loadfn(ptable_yaml_path)
    with open("ionic_radii.csv") as file:
        radii_data = file.read()
    radii_data = radii_data.split("\r")
    header = radii_data[0].split(",")
    for idx in range(1, len(radii_data)):
        line = radii_data[idx]
        tokens = line.strip().split(",")
        suffix = ""
        name = tokens[1]
        if len(name.split(" ")) > 1:
            suffix = "_" + name.split(" ")[1]
        el = tokens[2]

        ionic_radii = {}
        for tok_idx in range(3, len(tokens)):
            if match := re.match(r"^\s*([0-9\.]+)", tokens[tok_idx]):
                ionic_radii[int(header[tok_idx])] = float(match[1])

        if el in data:
            data[el][f"Ionic_radii{suffix}"] = ionic_radii
            if suffix == "_hs":
                data[el]["Ionic_radii"] = ionic_radii
        else:
            print(el)
    with open("periodic_table2.yaml", mode="w") as file:
        yaml.dump(data, file)


def parse_radii():
    data = loadfn(ptable_yaml_path)
    with open("radii.csv") as file:
        radii_data = file.read()
    radii_data = radii_data.split("\r")

    for line in radii_data:
        tokens = line.strip().split(",")
        el = tokens[1]
        try:
            atomic_radii = float(tokens[3]) / 100
        except Exception:
            atomic_radii = tokens[3]

        try:
            atomic_radii_calc = float(tokens[4]) / 100
        except Exception:
            atomic_radii_calc = tokens[4]

        try:
            vdw_radii = float(tokens[5]) / 100
        except Exception:
            vdw_radii = tokens[5]

        if el in data:
            data[el]["Atomic radius"] = atomic_radii
            data[el]["Atomic radius calculated"] = atomic_radii_calc
            data[el]["Van der waals radius"] = vdw_radii
        else:
            print(el)
    with open("periodic_table2.yaml", mode="w") as file:
        yaml.dump(data, file)
    with open("../pymatgen/core/periodic_table.json", mode="w") as file:
        json.dump(data, file)


def update_ionic_radii():
    data = loadfn(ptable_yaml_path)

    for dct in data.values():
        if "Ionic_radii" in dct:
            dct["Ionic radii"] = {k: v / 100 for k, v in dct["Ionic_radii"].items()}
            del dct["Ionic_radii"]
        if "Ionic_radii_hs" in dct:
            dct["Ionic radii hs"] = {k: v / 100 for k, v in dct["Ionic_radii_hs"].items()}
            del dct["Ionic_radii_hs"]
        if "Ionic_radii_ls" in dct:
            dct["Ionic radii ls"] = {k: v / 100 for k, v in dct["Ionic_radii_ls"].items()}
            del dct["Ionic_radii_ls"]
    with open("periodic_table2.yaml", mode="w") as file:
        yaml.dump(data, file)
    with open("../pymatgen/core/periodic_table.json", mode="w") as file:
        json.dump(data, file)


def parse_shannon_radii():
    data = loadfn(ptable_yaml_path)

    from openpyxl import load_workbook

    wb = load_workbook("Shannon Radii.xlsx")
    print(wb.sheetnames())
    sheet = wb["Sheet1"]
    i = 2
    el = charge = cn = None
    radii = defaultdict(dict)
    while sheet[f"E{i}"].value:
        if sheet[f"A{i}"].value:
            el = sheet[f"A{i}"].value
        if sheet[f"B{i}"].value:
            charge = int(sheet[f"B{i}"].value)
            radii[el][charge] = {}
        if sheet[f"C{i}"].value:
            cn = sheet[f"C{i}"].value
            radii[el][charge].setdefault(cn, {})

        spin = sheet[f"D{i}"].value if sheet[f"D{i}"].value is not None else ""

        radii[el][charge][cn][spin] = {
            "crystal_radius": float(sheet[f"E{i}"].value),
            "ionic_radius": float(sheet[f"F{i}"].value),
        }
        i += 1

    for el in radii:
        if el in data:
            data[el]["Shannon radii"] = dict(radii[el])

    dumpfn(data, ptable_yaml_path)
    with open("../pymatgen/core/periodic_table.json", mode="w") as file:
        json.dump(data, file)


def gen_periodic_table():
    data = loadfn(ptable_yaml_path)

    with open("../pymatgen/core/periodic_table.json", mode="w") as file:
        json.dump(data, file)


def gen_iupac_ordering():
    periodic_table = loadfn("../pymatgen/core/periodic_table.json")
    order = [
        ([18], range(6, 0, -1)),  # noble gasses
        ([1], range(7, 1, -1)),  # alkali metals
        ([2], range(7, 1, -1)),  # alkali earth metals
        (range(17, 2, -1), [9]),  # actinides
        (range(17, 2, -1), [8]),  # lanthanides
        ([3], (5, 4)),  # Y, Sc
        ([4], (6, 5, 4)),  # Hf -> Ti
        ([5], (6, 5, 4)),  # Ta -> V
        ([6], (6, 5, 4)),  # W -> Cr
        ([7], (6, 5, 4)),  # Re -> Mn
        ([8], (6, 5, 4)),  # Os -> Fe
        ([9], (6, 5, 4)),  # Ir -> Co
        ([10], (6, 5, 4)),  # Pt -> Ni
        ([11], (6, 5, 4)),  # Au -> Cu
        ([12], (6, 5, 4)),  # Hg -> Zn
        ([13], range(6, 1, -1)),  # Tl -> B
        ([14], range(6, 1, -1)),  # Pb -> C
        ([15], range(6, 1, -1)),  # Bi -> N
        ([1], [1]),  # Hydrogen
        ([16], range(6, 1, -1)),  # Po -> O
        ([17], range(6, 1, -1)),
    ]  # At -> F

    order = sum((list(product(x, y)) for x, y in order), [])  # noqa: RUF017
    iupac_ordering_dict = dict(zip([Element.from_row_and_group(row, group) for group, row in order], range(len(order))))

    # first clean periodic table of any IUPAC ordering
    for el in periodic_table:
        periodic_table[el].pop("IUPAC ordering", None)

    # now add iupac ordering
    for el in periodic_table:
        if "IUPAC ordering" in periodic_table[el]:
            # sanity check that we don't cover the same element twice
            raise KeyError(f"IUPAC ordering already exists for {el}")

        periodic_table[el]["IUPAC ordering"] = iupac_ordering_dict[get_el_sp(el)]


@requires(BeautifulSoup, "BeautifulSoup must be installed to use this method.")
def add_electron_affinities():
    """Update the periodic table data file with electron affinities."""

    req = requests.get("https://wikipedia.org/wiki/Electron_affinity_(data_page)", timeout=600)
    soup = BeautifulSoup(req.text, "html.parser")
    table = None
    for table in soup.find_all("table"):
        if "Hydrogen" in table.text:
            break
    data = []
    for tr in table.find_all("tr"):
        row = []
        for td in tr.find_all("td"):
            row += [td.get_text().strip()]
        data += [row]
    data.pop(0)

    ea = {}
    max_Z = max(Element(element).Z for element in Element.__members__)
    for r in data:
        # don't want superheavy elements or less common isotopes
        if int(r[0]) > max_Z or r[2] in ea:
            continue
        temp_str = re.sub(r"[\s\(\)]", "", r[3].strip("()[]"))
        # hyphen-like characters used that can't be parsed by .float
        bytes_rep = temp_str.encode("unicode_escape").replace(b"\\u2212", b"-")
        ea[r[2]] = float(bytes_rep.decode("unicode_escape"))

    Z_set = {Element.from_name(element).Z for element in ea}
    assert Z_set.issuperset(range(1, 93))  # Ensure that we have data for up to U.
    print(ea)
    pt = loadfn("../pymatgen/core/periodic_table.json")
    for key, val in pt.items():
        val["Electron affinity"] = ea.get(Element(key).long_name)
    dumpfn(pt, "../pymatgen/core/periodic_table.json")


def add_ionization_energies():
    """Update the periodic table data file with ground level and ionization energies from NIST."""

    with open("NIST Atomic Ionization Energies Output.html") as file:
        soup = BeautifulSoup(file.read(), "html.parser")
    table = None
    for table in soup.find_all("table"):
        if "Hydrogen" in table.text:
            break
    data = defaultdict(list)
    for row in table.find_all("tr"):
        row = [td.get_text().strip() for td in row.find_all("td")]
        if row:
            Z = int(row[0])
            val = re.sub(r"\s", "", row[8].strip("()[]"))
            val = None if val == "" else float(val)
            data[Z] += [val]
    print(data)
    print(data[51])
    assert set(data).issuperset(range(1, 93))  # Ensure that we have data for up to U.
    pt = loadfn("../pymatgen/core/periodic_table.json")
    for key, val in pt.items():
        del val["Ionization energy"]
        val["Ionization energies"] = data.get(Element(key).long_name, [])
    dumpfn(pt, "../pymatgen/core/periodic_table.json")


if __name__ == "__main__":
    # parse_shannon_radii()
    # add_ionization_energies()
    add_electron_affinities()
    # gen_periodic_table()
